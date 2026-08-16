from django.shortcuts import render, get_object_or_404
from django.contrib.auth.decorators import login_required
from django.http import HttpResponse
from accounts.models import Company
from ..models import Employee, Payslip

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from reportlab.lib.pagesizes import letter, landscape
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.lib import colors


MONTHS_LIST = [
    "Janvier", "Février", "Mars", "Avril", "Mai", "Juin",
    "Juillet", "Août", "Septembre", "Octobre", "Novembre", "Décembre"
]


def _parse_period_params(request):
    """Extrait et sécurise les paramètres mois et année depuis la requête GET."""
    selected_month = request.GET.get('month', 'Janvier')
    try:
        selected_year = int(request.GET.get('year', 2026))
    except (ValueError, TypeError):
        selected_year = 2026
    return selected_month, selected_year


def _get_monthly_summary_data(company, selected_month, selected_year):
    """
    Construit la liste summary_data en croisant les employés
    et leurs bulletins de paie pour la période donnée.
    """
    # Retrait de is_active=True pour éviter la FieldError
    employees = Employee.objects.filter(company=company).order_by('last_name', 'first_name')
    payslips = Payslip.objects.filter(
        employee__company=company,
        month=selected_month,
        year=selected_year,
        is_archived=False
    ).select_related('employee')

    payslip_map = {p.employee_id: p for p in payslips}

    summary_data = []
    total_gross = 0
    total_cnaps = 0
    total_irsa = 0
    total_net = 0

    for emp in employees:
        p = payslip_map.get(emp.id)
        has_payslip = p is not None

        gross = p.gross_salary if has_payslip else 0
        cnaps = p.cnaps_employee if has_payslip else 0
        irsa = p.irsa if has_payslip else 0
        net = p.net_payable if has_payslip else 0

        summary_data.append({
            'employee': emp,
            'has_payslip': has_payslip,
            'gross_salary': gross,
            'cnaps_employee': cnaps,
            'irsa': irsa,
            'net_payable': net,
        })

        total_gross += gross
        total_cnaps += cnaps
        total_irsa += irsa
        total_net += net

    return {
        'summary_data': summary_data,
        'total_gross': total_gross,
        'total_cnaps': total_cnaps,
        'total_irsa': total_irsa,
        'total_net': total_net,
    }


@login_required
def monthly_summary_view(request):
    company = get_object_or_404(Company, owner=request.user)
    selected_month, selected_year = _parse_period_params(request)

    data = _get_monthly_summary_data(company, selected_month, selected_year)

    context = {
        'company': company,
        'months_list': MONTHS_LIST,
        'selected_month': selected_month,
        'selected_year': selected_year,
        **data,
    }
    return render(request, 'payroll/monthly_summary.html', context)


@login_required
def export_monthly_summary_excel(request):
    company = get_object_or_404(Company, owner=request.user)
    selected_month, selected_year = _parse_period_params(request)

    data = _get_monthly_summary_data(company, selected_month, selected_year)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f"Paie {selected_month} {selected_year}"

    header_font = Font(name='Calibri', size=11, bold=True, color='FFFFFF')
    header_fill = PatternFill(start_color='212529', end_color='212529', fill_type='solid')
    bold_font = Font(name='Calibri', size=11, bold=True)
    right_align = Alignment(horizontal='right', vertical='center')

    ws.merge_cells('A1:G1')
    ws['A1'] = f"JOURNAL RECAPITULATIF DE PAIE - {selected_month.upper()} {selected_year} ({company.name})"
    ws['A1'].font = Font(size=14, bold=True)

    headers = ["Matricule", "Nom & Prénoms", "Salaire Brut", "CNaPS (1%)", "IRSA", "Net à Payer", "Statut"]
    ws.append([])
    ws.append(headers)

    for col in range(1, 8):
        cell = ws.cell(row=3, column=col)
        cell.font = header_font
        cell.fill = header_fill

    for item in data['summary_data']:
        emp = item['employee']
        matricule = getattr(emp, 'registration_number', None) or str(emp.id)
        statut = "Généré" if item['has_payslip'] else "Non généré"

        ws.append([
            matricule,
            f"{emp.last_name} {emp.first_name}",
            float(item['gross_salary']),
            float(item['cnaps_employee']),
            float(item['irsa']),
            float(item['net_payable']),
            statut
        ])
        row_idx = ws.max_row
        for c in range(3, 7):
            ws.cell(row=row_idx, column=c).number_format = '#,##0.00 "Ar"'
            ws.cell(row=row_idx, column=c).alignment = right_align

    ws.append([
        "TOTAL GENERAL", "",
        float(data['total_gross']),
        float(data['total_cnaps']),
        float(data['total_irsa']),
        float(data['total_net']),
        ""
    ])
    last_row = ws.max_row
    for col in range(1, 8):
        cell = ws.cell(row=last_row, column=col)
        cell.font = bold_font
        if 3 <= col <= 6:
            cell.number_format = '#,##0.00 "Ar"'
            cell.alignment = right_align

    ws.column_dimensions['A'].width = 15
    ws.column_dimensions['B'].width = 30
    for col in ['C', 'D', 'E', 'F']:
        ws.column_dimensions[col].width = 18
    ws.column_dimensions['G'].width = 15

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="recap_paie_{selected_month}_{selected_year}.xlsx"'
    wb.save(response)
    return response


@login_required
def export_monthly_summary_pdf(request):
    company = get_object_or_404(Company, owner=request.user)
    selected_month, selected_year = _parse_period_params(request)

    data = _get_monthly_summary_data(company, selected_month, selected_year)

    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = f'inline; filename="recap_paie_{selected_month}_{selected_year}.pdf"'

    doc = SimpleDocTemplate(response, pagesize=landscape(letter), rightMargin=30, leftMargin=30, topMargin=30, bottomMargin=30)
    elements = []
    styles = getSampleStyleSheet()

    elements.append(Paragraph(f"<b>RECAPITULATIF MENSUEL DE PAIE - {selected_month.upper()} {selected_year}</b>", styles['Heading1']))
    elements.append(Paragraph(f"Société : {company.name}", styles['Normal']))
    elements.append(Spacer(1, 15))

    table_data = [["Matricule", "Nom & Prénoms", "Salaire Brut", "CNaPS (1%)", "IRSA", "Net à Payer", "Statut"]]
    for item in data['summary_data']:
        emp = item['employee']
        matricule = getattr(emp, 'registration_number', None) or str(emp.id)
        statut = "Généré" if item['has_payslip'] else "Non généré"

        table_data.append([
            matricule,
            f"{emp.last_name} {emp.first_name}",
            f"{item['gross_salary']:,.2f} Ar",
            f"{item['cnaps_employee']:,.2f} Ar",
            f"{item['irsa']:,.2f} Ar",
            f"{item['net_payable']:,.2f} Ar",
            statut
        ])

    table_data.append([
        "TOTAL", "GENERAL",
        f"{data['total_gross']:,.2f} Ar",
        f"{data['total_cnaps']:,.2f} Ar",
        f"{data['total_irsa']:,.2f} Ar",
        f"{data['total_net']:,.2f} Ar",
        ""
    ])

    table = Table(table_data, colWidths=[80, 180, 100, 100, 100, 100, 80])
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor("#212529")),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (1, -1), 'LEFT'),
        ('ALIGN', (2, 0), (5, -1), 'RIGHT'),
        ('ALIGN', (6, 0), (6, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor("#DEE2E6")),
        ('FONTNAME', (0, -1), (-1, -1), 'Helvetica-Bold'),
        ('BACKGROUND', (0, -1), (-1, -1), colors.HexColor("#E2E3E5")),
    ]))

    elements.append(table)
    doc.build(elements)
    return response