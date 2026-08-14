from io import BytesIO
import re
from urllib.parse import quote
from datetime import datetime

from django.db.models import Q
from django.http import HttpResponse, HttpResponseForbidden
from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required
from django.contrib import messages

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

from reportlab.lib.pagesizes import A4, letter, landscape
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib import colors

from accounts.models import Company
from .models import Employee, Payslip
from .forms import EmployeeForm, PayslipCreateForm
from .utils.calculations import calculate_madagascar_payroll


# =====================================================================
# DASHBOARD
# =====================================================================

@login_required
def dashboard(request):
    company = Company.objects.filter(owner=request.user).first()
    if not company:
        return redirect('accounts:company_create')

    employees = Employee.objects.filter(company=company)
    payslips_queryset = Payslip.objects.filter(employee__company=company, is_archived=False)

    search_query = request.GET.get('q', '').strip()
    if search_query:
        payslips_queryset = payslips_queryset.filter(
            Q(employee__first_name__icontains=search_query) |
            Q(employee__last_name__icontains=search_query) |
            Q(month__icontains=search_query) |
            Q(year__icontains=search_query)
        )

    recent_payslips = payslips_queryset.order_by('-created_at')

    context = {
        'company': company,
        'employees': employees,
        'recent_payslips': recent_payslips,
        'search_query': search_query,
    }
    return render(request, 'payroll/dashboard.html', context)


# =====================================================================
# GESTION DES SALARIÉS (CRUD)
# =====================================================================

@login_required
def employee_list(request):
    company = get_object_or_404(Company, owner=request.user)
    employees = Employee.objects.filter(company=company)
    return render(request, 'payroll/employee_list.html', {'employees': employees})


@login_required
def employee_add(request):
    company = get_object_or_404(Company, owner=request.user)
    if request.method == 'POST':
        form = EmployeeForm(request.POST)
        if form.is_valid():
            employee = form.save(commit=False)
            employee.company = company
            employee.save()
            messages.success(request, f"Le salarié {employee.first_name} {employee.last_name} a été ajouté.")
            return redirect('payroll:dashboard')
    else:
        form = EmployeeForm()
    return render(request, 'payroll/employee_form.html', {'form': form})


@login_required
def employee_detail(request, employee_id):
    employee = get_object_or_404(Employee, id=employee_id, company__owner=request.user)
    payslips = employee.payslips.filter(is_archived=False).order_by('-created_at')
    return render(request, 'payroll/employee_detail.html', {
        'employee': employee,
        'payslips': payslips
    })


@login_required
def employee_edit(request, employee_id):
    employee = get_object_or_404(Employee, id=employee_id, company__owner=request.user)
    if request.method == 'POST':
        form = EmployeeForm(request.POST, instance=employee)
        if form.is_valid():
            form.save()
            messages.success(request, "Les informations du salarié ont été mises à jour.")
            return redirect('payroll:dashboard')
    else:
        form = EmployeeForm(instance=employee)
    return render(request, 'payroll/employee_form.html', {'form': form, 'employee': employee})


@login_required
def employee_delete(request, employee_id):
    employee = get_object_or_404(Employee, id=employee_id, company__owner=request.user)
    if request.method == 'POST':
        employee.delete()
        messages.success(request, "Salarié supprimé avec succès.")
        return redirect('payroll:dashboard')
    return render(request, 'payroll/employee_confirm_delete.html', {'employee': employee})


# =====================================================================
# GESTION DES BULLETINS DE PAIE
# =====================================================================

@login_required
def payslip_list(request):
    company = get_object_or_404(Company, owner=request.user)
    payslips = Payslip.objects.filter(employee__company=company, is_archived=False).order_by('-created_at')
    return render(request, 'payroll/payslip_list.html', {'payslips': payslips})


@login_required
def payslip_detail(request, payslip_id):
    payslip = get_object_or_404(Payslip, id=payslip_id, employee__company__owner=request.user)
    return render(request, 'payroll/payslip_detail.html', {'payslip': payslip})


@login_required
def payslip_create(request, employee_id):
    employee = get_object_or_404(Employee, id=employee_id, company__owner=request.user)
    if request.method == 'POST':
        form = PayslipCreateForm(request.POST)
        if form.is_valid():
            payslip = form.save(commit=False)
            payslip.employee = employee

            # 1. Retenue pour absence
            payslip.unpaid_absences_amount = (payslip.absence_days or 0) * (payslip.daily_absence_rate or 0)

            # 2. Calculs automatiques des cotisations et de l'IRSA
            calc = calculate_madagascar_payroll(
                base_salary=employee.base_salary,
                unpaid_absences=payslip.unpaid_absences_amount,
                overtime=payslip.overtime_amount or 0,
                bonus=payslip.bonuses or 0,
                advance=payslip.advances or 0,
                children=employee.children_count or 0
            )
            payslip.gross_salary = calc['gross_salary']
            payslip.cnaps_employee = calc['cnaps_employee']
            payslip.smids_employee = calc['smids_employee']
            payslip.taxable_base = calc['taxable_base']
            payslip.irsa = calc['irsa']
            payslip.net_payable = calc['net_payable']
            payslip.cnaps_employer = calc['cnaps_employer']
            payslip.smids_employer = calc['smids_employer']
            payslip.fmfp_employer = calc['fmfp_employer']
            payslip.save()

            # 3. Mise à jour du solde de l'avance du salarié
            if payslip.advances > 0 and employee.total_advance_balance > 0:
                employee.total_advance_balance = max(0, employee.total_advance_balance - payslip.advances)

            # 4. Mise à jour du solde de congés payés restants
            paid_leave = getattr(payslip, 'paid_leave_days', 0) or 0
            if paid_leave > 0 and getattr(employee, 'leave_balance', 0) > 0:
                employee.leave_balance = max(0, employee.leave_balance - paid_leave)

            employee.save()

            # 5. INCRÉMENTATION DU COMPTEUR DE GÉNÉRATIONS
            company = employee.company
            if company:
                company.generation_count += 1
                company.save()

            messages.success(request, f"Bulletin généré avec succès pour {employee.first_name} {employee.last_name}.")
            return redirect('payroll:payslip_detail', payslip_id=payslip.id)
        else:
            messages.error(request, "Veuillez corriger les erreurs ci-dessous.")
    else:
        default_daily_rate = round(employee.base_salary / 30, 2) if employee.base_salary else 0
        form = PayslipCreateForm(initial={
            'year': 2026,
            'working_days': 30,
            'daily_absence_rate': default_daily_rate
        })
    return render(request, 'payroll/payslip_form.html', {'form': form, 'employee': employee})


@login_required
def payslip_bulk_create(request):
    """Génération groupée des fiches de paie pour tous les salariés de l'entreprise"""
    company = get_object_or_404(Company, owner=request.user)
    employees = Employee.objects.filter(company=company)

    if request.method == 'POST':
        month = request.POST.get('month')
        year = int(request.POST.get('year', 2026))

        if not month:
            messages.error(request, "Veuillez sélectionner un mois.")
            return redirect('payroll:payslip_bulk_create')

        count = 0
        for employee in employees:
            default_daily_rate = round(employee.base_salary / 30, 2) if employee.base_salary else 0
            calc = calculate_madagascar_payroll(
                base_salary=employee.base_salary,
                unpaid_absences=0,
                overtime=0,
                bonus=0,
                advance=0,
                children=employee.children_count or 0
            )
            Payslip.objects.create(
                employee=employee,
                month=month,
                year=year,
                working_days=30,
                absence_days=0,
                paid_leave_days=0,
                daily_absence_rate=default_daily_rate,
                unpaid_absences_amount=0,
                overtime_hours=0,
                overtime_amount=0,
                bonuses=0,
                advances=0,
                gross_salary=calc['gross_salary'],
                cnaps_employee=calc['cnaps_employee'],
                smids_employee=calc['smids_employee'],
                taxable_base=calc['taxable_base'],
                irsa=calc['irsa'],
                net_payable=calc['net_payable'],
                cnaps_employer=calc['cnaps_employer'],
                smids_employer=calc['smids_employer'],
                fmfp_employer=calc['fmfp_employer']
            )
            count += 1

        if count > 0:
            company.generation_count += count
            company.save()

        messages.success(request, f"{count} bulletin(s) généré(s) avec succès pour {month} {year}.")
        return redirect('payroll:dashboard')

    months_list = [
        "Janvier", "Février", "Mars", "Avril", "Mai", "Juin",
        "Juillet", "Août", "Septembre", "Octobre", "Novembre", "Décembre"
    ]

    context = {
        'employees': employees,
        'months_list': months_list,
        'current_year': 2026
    }
    return render(request, 'payroll/payslip_bulk_form.html', context)


@login_required
def toggle_hide_payslip(request, payslip_id):
    """Bascule le statut masqué / affiché d'un bulletin de paie"""
    payslip = get_object_or_404(Payslip, id=payslip_id, employee__company__owner=request.user)
    payslip.is_archived = not payslip.is_archived
    payslip.save()

    status_msg = "masqué" if payslip.is_archived else "restauré"
    messages.info(request, f"Le bulletin de paie a été {status_msg}.")
    return redirect('payroll:dashboard')


@login_required
def archived_payslips(request):
    """Affiche la liste des fiches de paie masquées / archivées"""
    company = get_object_or_404(Company, owner=request.user)
    archived_list = Payslip.objects.filter(employee__company=company, is_archived=True).order_by('-created_at')

    return render(request, 'payroll/archived_payslips.html', {
        'archived_payslips': archived_list
    })


# =====================================================================
# IMPRESSION PDF - BULLETIN DE PAIE INDIVIDUEL
# =====================================================================

@login_required
def generate_payslip_pdf(request, payslip_id):
    payslip = get_object_or_404(Payslip, id=payslip_id)

    if payslip.employee.company.owner != request.user:
        return HttpResponseForbidden("Vous n'avez pas l'autorisation d'accéder à ce document.")

    buffer = BytesIO()
    doc = SimpleDocTemplate(
        buffer,
        pagesize=A4,
        rightMargin=25,
        leftMargin=25,
        topMargin=20,
        bottomMargin=20
    )

    story = []
    styles = getSampleStyleSheet()

    title_style = ParagraphStyle(
        'TitleStyle', parent=styles['Heading1'],
        fontName='Helvetica-Bold', fontSize=12, alignment=1, textColor=colors.black
    )
    cell_style = ParagraphStyle('CellStyle', fontName='Helvetica', fontSize=8, leading=10)
    bold_style = ParagraphStyle('BoldStyle', fontName='Helvetica-Bold', fontSize=8, leading=10)
    right_style = ParagraphStyle('RightStyle', fontName='Helvetica', fontSize=8, leading=10, alignment=2)
    bold_right = ParagraphStyle('BoldRight', fontName='Helvetica-Bold', fontSize=8, leading=10, alignment=2)

    # 1. Titre
    title_data = [[Paragraph("BULLETIN DE PAIE", title_style)]]
    title_table = Table(title_data, colWidths=[545])
    title_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, -1), colors.HexColor("#EAEAEA")),
        ('BOX', (0, 0), (-1, -1), 1, colors.black),
        ('TOPPADDING', (0, 0), (-1, -1), 5),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 5),
    ]))
    story.append(title_table)
    story.append(Spacer(1, 8))

    # 2. Section Employeur
    emp = payslip.employee.company
    emp_data = [
        [
            Paragraph(f"<b>{emp.name}</b><br/>{emp.address}", cell_style),
            Paragraph(
                f"<b>NIF :</b> {emp.nif}<br/><b>STAT :</b> {getattr(emp, 'stat', None) or '-'}<br/><b>CNaPS :</b> {emp.cnaps_no}",
                cell_style)
        ]
    ]
    emp_table = Table(emp_data, colWidths=[272, 273])
    emp_table.setStyle(TableStyle([
        ('GRID', (0, 0), (-1, -1), 0.5, colors.black),
        ('VALIGN', (0, 0), (-1, -1), 'TOP'),
        ('PADDING', (0, 0), (-1, -1), 4),
    ]))
    story.append(emp_table)
    story.append(Spacer(1, 6))

    # 3. Section Salarié
    e = payslip.employee
    hire_date_str = e.hire_date.strftime("%d/%m/%Y") if getattr(e, 'hire_date', None) else "-"

    paid_leave = getattr(payslip, 'paid_leave_days', 0) or 0
    leave_bal = getattr(e, 'leave_balance', 0) or 0
    time_info_str = f"{payslip.working_days} J. | Congés pris: {paid_leave} j (Solde: {leave_bal} j) | Abs. non payées: {payslip.absence_days} j"

    employee_data = [
        [Paragraph("PAYE MOIS", bold_style), Paragraph(f"{payslip.month} {payslip.year}", cell_style),
         Paragraph("DATE D'EMBAUCHE", bold_style), Paragraph(hire_date_str, cell_style)],
        [Paragraph("NOM & PRENOM", bold_style), Paragraph(f"{e.last_name} {e.first_name}", cell_style),
         Paragraph("DEPARTEMENT", bold_style), Paragraph(str(getattr(e, 'department', None) or '-'), cell_style)],
        [Paragraph("CIN", bold_style), Paragraph(str(getattr(e, 'cin', None) or '-'), cell_style),
         Paragraph("CATEGORIE", bold_style), Paragraph(str(getattr(e, 'category', None) or '-'), cell_style)],
        [Paragraph("FONCTION", bold_style), Paragraph(str(getattr(e, 'job_title', None) or '-'), cell_style),
         Paragraph("CNaPS SALARIÉ", bold_style), Paragraph(str(getattr(e, 'cnaps_no', None) or '-'), cell_style)],
        [Paragraph("CONTRAT", bold_style), Paragraph(str(getattr(e, 'contract_type', None) or '-'), cell_style),
         Paragraph("ENFANTS À CHARGE", bold_style), Paragraph(str(getattr(e, 'children_count', 0)), cell_style)],
        [Paragraph("SALAIRE DE BASE", bold_style), Paragraph(f"<b>{e.base_salary:,.2f} Ar</b>", cell_style),
         Paragraph("TEMPS & CONGÉS", bold_style), Paragraph(time_info_str, cell_style)],
    ]
    emp_info_table = Table(employee_data, colWidths=[110, 162, 120, 153])
    emp_info_table.setStyle(TableStyle([
        ('GRID', (0, 0), (-1, -1), 0.5, colors.black),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('PADDING', (0, 0), (-1, -1), 3),
    ]))
    story.append(emp_info_table)
    story.append(Spacer(1, 8))

    # 4. Tableau Calculs de Paie
    absence_label = f"Absences ({payslip.absence_days} j × {payslip.daily_absence_rate:,.2f} Ar)" if payslip.absence_days > 0 else "Absences / Retenues"
    advance_label = f"Retenue Avance (Reste dû: {e.total_advance_balance:,.2f} Ar)" if e.total_advance_balance > 0 else "Avance sur salaire"

    calc_data = [
        [
            Paragraph("<b>DÉSIGNATION</b>", bold_style),
            Paragraph("<b>GAINS</b>", bold_right),
            Paragraph("<b>RETENUES</b>", bold_right),
            Paragraph("<b>TAUX PAT.</b>", bold_style),
            Paragraph("<b>MONTANT PAT.</b>", bold_right)
        ],
        [Paragraph("Salaire de Base", cell_style), Paragraph(f"{e.base_salary:,.2f}", right_style),
         Paragraph("", cell_style), Paragraph("", cell_style), Paragraph("", cell_style)],
        [Paragraph(absence_label, cell_style), Paragraph("", cell_style),
         Paragraph(f"{payslip.unpaid_absences_amount:,.2f}", right_style), Paragraph("", cell_style),
         Paragraph("", cell_style)],
        [Paragraph("Heures Supplémentaires", cell_style), Paragraph(f"{payslip.overtime_amount:,.2f}", right_style),
         Paragraph("", cell_style), Paragraph("", cell_style), Paragraph("", cell_style)],
        [Paragraph("Primes et Indemnités", cell_style), Paragraph(f"{payslip.bonuses:,.2f}", right_style),
         Paragraph("", cell_style), Paragraph("", cell_style), Paragraph("", cell_style)],
        [Paragraph("<b>TOTAL BRUT</b>", bold_style), Paragraph(f"<b>{payslip.gross_salary:,.2f}</b>", bold_right),
         Paragraph("", cell_style), Paragraph("", cell_style), Paragraph("", cell_style)],
        [Paragraph("CNaPS (1%)", cell_style), Paragraph("", cell_style),
         Paragraph(f"{payslip.cnaps_employee:,.2f}", right_style), Paragraph("13%", cell_style),
         Paragraph(f"{payslip.cnaps_employer:,.2f}", right_style)],
        [Paragraph("SMIDS / OSTIE", cell_style), Paragraph("", cell_style),
         Paragraph(f"{payslip.smids_employee:,.2f}", right_style), Paragraph("5%", cell_style),
         Paragraph(f"{payslip.smids_employer:,.2f}", right_style)],
        [Paragraph("FMFP", cell_style), Paragraph("", cell_style), Paragraph("", cell_style),
         Paragraph("1%", cell_style), Paragraph(f"{payslip.fmfp_employer:,.2f}", right_style)],
        [Paragraph("<b>BASE IMPOSABLE</b>", bold_style), Paragraph(f"<b>{payslip.taxable_base:,.2f}</b>", bold_right),
         Paragraph("", cell_style), Paragraph("", cell_style), Paragraph("", cell_style)],
        [Paragraph("<b>IRSA</b>", bold_style), Paragraph("", cell_style),
         Paragraph(f"<b>{payslip.irsa:,.2f}</b>", bold_right), Paragraph("", cell_style), Paragraph("", cell_style)],
        [Paragraph(advance_label, cell_style), Paragraph("", cell_style),
         Paragraph(f"{payslip.advances:,.2f}", right_style), Paragraph("", cell_style), Paragraph("", cell_style)],
        [Paragraph("<b>NET À PAYER</b>", bold_style), Paragraph(f"<b>{payslip.net_payable:,.2f} Ar</b>", bold_right),
         Paragraph("", cell_style), Paragraph("", cell_style), Paragraph("", cell_style)],
    ]

    calc_table = Table(calc_data, colWidths=[205, 85, 85, 70, 100])
    calc_table.setStyle(TableStyle([
        ('GRID', (0, 0), (-1, -1), 0.5, colors.black),
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor("#F0F0F0")),
        ('BACKGROUND', (0, 5), (1, 5), colors.HexColor("#F0F0F0")),
        ('BACKGROUND', (0, 9), (1, 9), colors.HexColor("#F0F0F0")),
        ('BACKGROUND', (0, 12), (1, 12), colors.HexColor("#EAEAEA")),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('PADDING', (0, 0), (-1, -1), 3),
    ]))
    story.append(calc_table)
    story.append(Spacer(1, 15))

    # 5. Signatures
    sig_data = [
        [Paragraph("<b><u>L'EMPLOYEUR</u></b>", bold_style), Paragraph("<b><u>LE SALARIÉ</u></b>", bold_style)]
    ]
    sig_table = Table(sig_data, colWidths=[272, 273])
    sig_table.setStyle(TableStyle([
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('VALIGN', (0, 0), (-1, -1), 'TOP'),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 30),
    ]))
    story.append(sig_table)

    doc.build(story)

    pdf_data = buffer.getvalue()
    buffer.close()

    response = HttpResponse(pdf_data, content_type='application/pdf')

    last_name = re.sub(r'[^\w\s-]', '', e.last_name).strip().replace(' ', '_')
    first_name = re.sub(r'[^\w\s-]', '', e.first_name).strip().replace(' ', '_')
    month = re.sub(r'[^\w\s-]', '', str(payslip.month)).strip().replace(' ', '_')
    year = str(payslip.year)

    filename = f"Bulletin_{last_name}_{first_name}_{month}_{year}.pdf"
    encoded_filename = quote(filename)

    response['Content-Disposition'] = f'attachment; filename="{filename}"; filename*=UTF-8\'\'{encoded_filename}'
    return response


# =====================================================================
# DÉCLARATIONS SOCIALES, FISCALES ET RÉCAPITULATIF MENSUEL
# =====================================================================

@login_required
def irsa_declaration_view(request):
    """Génère l'état récapitulatif IRSA pour un mois et une année donnés"""
    company = get_object_or_404(Company, owner=request.user)

    months_list = [
        "Janvier", "Février", "Mars", "Avril", "Mai", "Juin",
        "Juillet", "Août", "Septembre", "Octobre", "Novembre", "Décembre"
    ]

    selected_month = request.GET.get('month', "Janvier")
    selected_year = int(request.GET.get('year', 2026))

    payslips = Payslip.objects.filter(
        employee__company=company,
        month=selected_month,
        year=selected_year,
        is_archived=False
    )

    total_gross = sum(p.gross_salary for p in payslips)
    total_taxable_base = sum(p.taxable_base for p in payslips)
    total_irsa = sum(p.irsa for p in payslips)

    context = {
        'company': company,
        'payslips': payslips,
        'months_list': months_list,
        'selected_month': selected_month,
        'selected_year': selected_year,
        'total_gross': total_gross,
        'total_taxable_base': total_taxable_base,
        'total_irsa': total_irsa,
    }
    return render(request, 'payroll/irsa_declaration.html', context)


@login_required
def cnaps_declaration_view(request):
    """Génère le bordereau nominatif CNaPS pour un mois et une année donnés"""
    company = get_object_or_404(Company, owner=request.user)

    months_list = [
        "Janvier", "Février", "Mars", "Avril", "Mai", "Juin",
        "Juillet", "Août", "Septembre", "Octobre", "Novembre", "Décembre"
    ]

    selected_month = request.GET.get('month', "Janvier")
    selected_year = int(request.GET.get('year', 2026))

    payslips = Payslip.objects.filter(
        employee__company=company,
        month=selected_month,
        year=selected_year,
        is_archived=False
    )

    total_cnaps_employee = sum(p.cnaps_employee for p in payslips)
    total_cnaps_employer = sum(p.cnaps_employer for p in payslips)
    total_cnaps_due = total_cnaps_employee + total_cnaps_employer

    context = {
        'company': company,
        'payslips': payslips,
        'months_list': months_list,
        'selected_month': selected_month,
        'selected_year': selected_year,
        'total_cnaps_employee': total_cnaps_employee,
        'total_cnaps_employer': total_cnaps_employer,
        'total_cnaps_due': total_cnaps_due,
    }
    return render(request, 'payroll/cnaps_declaration.html', context)


@login_required
def health_declaration_view(request):
    """Génère l'état récapitulatif pour la Médecine d'Entreprise"""
    company = get_object_or_404(Company, owner=request.user)

    months_list = [
        "Janvier", "Février", "Mars", "Avril", "Mai", "Juin",
        "Juillet", "Août", "Septembre", "Octobre", "Novembre", "Décembre"
    ]

    selected_month = request.GET.get('month', "Janvier")
    selected_year = int(request.GET.get('year', 2026))

    payslips = Payslip.objects.filter(
        employee__company=company,
        month=selected_month,
        year=selected_year,
        is_archived=False
    )

    total_smids_employee = sum(p.smids_employee for p in payslips)
    total_smids_employer = sum(p.smids_employer for p in payslips)
    total_health_due = total_smids_employee + total_smids_employer

    context = {
        'company': company,
        'payslips': payslips,
        'months_list': months_list,
        'selected_month': selected_month,
        'selected_year': selected_year,
        'total_smids_employee': total_smids_employee,
        'total_smids_employer': total_smids_employer,
        'total_health_due': total_health_due,
    }
    return render(request, 'payroll/health_declaration.html', context)


@login_required
def monthly_summary_view(request):
    """Affiche le récapitulatif mensuel de tous les salariés de l'entreprise (Journal de paie)"""
    company = get_object_or_404(Company, owner=request.user)

    months_list = [
        "Janvier", "Février", "Mars", "Avril", "Mai", "Juin",
        "Juillet", "Août", "Septembre", "Octobre", "Novembre", "Décembre"
    ]

    selected_month = request.GET.get('month', "Janvier")
    selected_year = int(request.GET.get('year', 2026))

    # 1. Tous les salariés de l'entreprise
    all_employees = Employee.objects.filter(company=company)

    # 2. Bulletins du mois sélectionné
    existing_payslips = Payslip.objects.filter(
        employee__company=company,
        month=selected_month,
        year=selected_year,
        is_archived=False
    )

    payslips_dict = {p.employee_id: p for p in existing_payslips}

    # 3. Construction du récapitulatif
    summary_data = []
    total_gross = 0
    total_cnaps = 0
    total_irsa = 0
    total_net = 0

    for emp in all_employees:
        payslip = payslips_dict.get(emp.id)
        if payslip:
            gross = payslip.gross_salary
            cnaps = payslip.cnaps_employee
            irsa = payslip.irsa
            net = payslip.net_payable
            has_payslip = True
        else:
            gross = 0
            cnaps = 0
            irsa = 0
            net = 0
            has_payslip = False

        total_gross += gross
        total_cnaps += cnaps
        total_irsa += irsa
        total_net += net

        summary_data.append({
            'employee': emp,
            'gross_salary': gross,
            'cnaps_employee': cnaps,
            'irsa': irsa,
            'net_payable': net,
            'has_payslip': has_payslip,
        })

    context = {
        'company': company,
        'summary_data': summary_data,
        'months_list': months_list,
        'selected_month': selected_month,
        'selected_year': selected_year,
        'total_gross': total_gross,
        'total_cnaps': total_cnaps,
        'total_irsa': total_irsa,
        'total_net': total_net,
    }
    return render(request, 'payroll/monthly_summary.html', context)


# =====================================================================
# EXPORTATIONS (EXCEL & PDF) - RÉCAPITULATIF MENSUEL
# =====================================================================

@login_required
def export_monthly_summary_excel(request):
    """Exporte le journal de paie mensuel au format Excel (.xlsx)"""
    company = get_object_or_404(Company, owner=request.user)
    month = request.GET.get('month', 'Janvier')
    year = int(request.GET.get('year', 2026))

    all_employees = Employee.objects.filter(company=company)
    existing_payslips = Payslip.objects.filter(
        employee__company=company, month=month, year=year, is_archived=False
    )
    payslips_dict = {p.employee_id: p for p in existing_payslips}

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f"Journal Paie {month} {year}"

    # En-têtes document
    ws.append([f"JOURNAL DE PAIE - {company.name.upper()}"])
    ws.append([f"Période : {month} {year}"])
    ws.append([])

    # En-têtes du tableau
    headers = ["Matricule", "Nom & Prénoms", "Salaire Brut (Ar)", "Retenue CNaPS 1% (Ar)", "Retenue IRSA (Ar)", "Net à Payer (Ar)"]
    ws.append(headers)

    header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True, size=11)

    for col in range(1, 7):
        cell = ws.cell(row=4, column=col)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center" if col <= 2 else "right", vertical="center")

    total_gross = total_cnaps = total_irsa = total_net = 0
    row_idx = 5

    for emp in all_employees:
        p = payslips_dict.get(emp.id)
        gross = p.gross_salary if p else 0
        cnaps = p.cnaps_employee if p else 0
        irsa = p.irsa if p else 0
        net = p.net_payable if p else 0

        total_gross += gross
        total_cnaps += cnaps
        total_irsa += irsa
        total_net += net

        reg_num = getattr(emp, 'registration_number', None) or str(emp.id)
        full_name = f"{emp.last_name} {emp.first_name}"

        ws.append([reg_num, full_name, gross, cnaps, irsa, net])

        for c_idx in range(3, 7):
            ws.cell(row=row_idx, column=c_idx).number_format = '#,##0.00'
        row_idx += 1

    # Ligne des Totaux
    ws.append(["TOTAL GÉNÉRAL", "", total_gross, total_cnaps, total_irsa, total_net])
    ws.merge_cells(start_row=row_idx, start_column=1, end_row=row_idx, end_column=2)

    total_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
    total_font = Font(bold=True)

    for col in range(1, 7):
        cell = ws.cell(row=row_idx, column=col)
        cell.fill = total_fill
        cell.font = total_font
        if col >= 3:
            cell.number_format = '#,##0.00'

    for col in ws.columns:
        max_len = max(len(str(cell.value or '')) for cell in col)
        col_letter = openpyxl.utils.get_column_letter(col[0].column)
        ws.column_dimensions[col_letter].width = max(max_len + 3, 14)

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="journal_paie_{month}_{year}.xlsx"'
    wb.save(response)
    return response


@login_required
def export_monthly_summary_pdf(request):
    """Exporte le journal de paie mensuel au format PDF (Paysage)"""
    company = get_object_or_404(Company, owner=request.user)
    month = request.GET.get('month', 'Janvier')
    year = int(request.GET.get('year', 2026))

    all_employees = Employee.objects.filter(company=company)
    existing_payslips = Payslip.objects.filter(
        employee__company=company, month=month, year=year, is_archived=False
    )
    payslips_dict = {p.employee_id: p for p in existing_payslips}

    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = f'inline; filename="journal_paie_{month}_{year}.pdf"'

    doc = SimpleDocTemplate(response, pagesize=landscape(letter), rightMargin=30, leftMargin=30, topMargin=30, bottomMargin=30)
    elements = []
    styles = getSampleStyleSheet()

    title_style = ParagraphStyle('TitleStyle', parent=styles['Heading1'], fontSize=16, leading=20, textColor=colors.HexColor('#1F4E78'))
    subtitle_style = ParagraphStyle('SubTitleStyle', parent=styles['Normal'], fontSize=11, leading=14, textColor=colors.gray)

    elements.append(Paragraph(f"<b>RÉCAPITULATIF MENSUEL DE PAIE (JOURNAL DE PAIE)</b>", title_style))
    elements.append(Paragraph(f"Société : <b>{company.name}</b> | Période : <b>{month} {year}</b>", subtitle_style))
    elements.append(Spacer(1, 15))

    data = [["Matricule", "Nom & Prénoms", "Salaire Brut", "CNaPS (1%)", "IRSA", "Net à Payer"]]

    total_gross = total_cnaps = total_irsa = total_net = 0

    for emp in all_employees:
        p = payslips_dict.get(emp.id)
        gross = p.gross_salary if p else 0
        cnaps = p.cnaps_employee if p else 0
        irsa = p.irsa if p else 0
        net = p.net_payable if p else 0

        total_gross += gross
        total_cnaps += cnaps
        total_irsa += irsa
        total_net += net

        reg_num = getattr(emp, 'registration_number', None) or str(emp.id)
        full_name = f"{emp.last_name} {emp.first_name}"

        data.append([
            reg_num,
            full_name,
            f"{gross:,.2f} Ar",
            f"{cnaps:,.2f} Ar",
            f"{irsa:,.2f} Ar",
            f"{net:,.2f} Ar"
        ])

    data.append([
        "TOTAL GÉNÉRAL",
        "",
        f"{total_gross:,.2f} Ar",
        f"{total_cnaps:,.2f} Ar",
        f"{total_irsa:,.2f} Ar",
        f"{total_net:,.2f} Ar"
    ])

    table = Table(data, colWidths=[90, 220, 110, 110, 110, 110])
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1F4E78')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 10),
        ('ALIGN', (2, 0), (-1, -1), 'RIGHT'),
        ('ALIGN', (0, 0), (1, -1), 'LEFT'),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
        ('TOPPADDING', (0, 0), (-1, -1), 6),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#D3D3D3')),
        ('BACKGROUND', (0, -1), (-1, -1), colors.HexColor('#D9E1F2')),
        ('FONTNAME', (0, -1), (-1, -1), 'Helvetica-Bold'),
        ('SPAN', (0, -1), (1, -1)),
    ]))

    elements.append(table)
    doc.build(elements)
    return response